import flet as ft
import openpyxl # Manipular Excel
import os # Acessa documentos através do sistema operacinal
from datetime import datetime # Capturar informações de data e hora
import pandas as pd


data_inicial_datetime = None
data_final_datetime = None


def main(page: ft.Page):

    
    cor_textos = '#F5F5F5'
    preto = '#3D3D3D'
    azul = "#4895EF"
    verde = "#75975e"
    grafite = '#747169'
    vermelho = '#ee6b6e'

    desc_porc_real = ft.Container(
        expand=True,
        padding=10,
        margin=5,
        border_radius=10,
        content=ft.Column([]),
    )

    total_entrada = ft.Container(
        height=60,
        width=120,
        content=ft.Row(
            [
                ft.Icon(name=ft.icons.NORTH, color=verde),
                ft.Text(value=0, size=18, weight=ft.FontWeight.BOLD, color=cor_textos),
            ],
            alignment=ft.MainAxisAlignment.CENTER,
        )
    )

    total_saida = ft.Container(
        height=60,
        width=120,
        content=ft.Row(
            [
                ft.Icon(name=ft.icons.SOUTH, color=vermelho),
                ft.Text(value=0, size=18, weight=ft.FontWeight.BOLD, color=cor_textos),
            ],
            alignment=ft.MainAxisAlignment.CENTER,
        )
    )

    saldo_total = ft.Container(
        margin=15,
        expand=True,
        border_radius=5,
        content=ft.Row(
            [
                ft.Text(value=0, size=30, weight=ft.FontWeight.BOLD, color=cor_textos)
            ],
            alignment=ft.MainAxisAlignment.END
        )
    )

    tipo = ft.Dropdown(
        bgcolor=verde,
        border_width=1,
        border_radius=15,
        color="#007B83",
        label_style=ft.TextStyle(size=15, color=cor_textos, weight=ft.FontWeight.BOLD),
        label='Tipo de Transação',
        options=
        [
            ft.dropdown.Option('Entrada', text_style=ft.TextStyle(size=15, color=cor_textos, weight=ft.FontWeight.BOLD)),
            ft.dropdown.Option('Saída', text_style=ft.TextStyle(size=15, color=cor_textos, weight=ft.FontWeight.BOLD)),                    
        ],
    )

    categoria = ft.Dropdown(
        bgcolor=verde,
        border_width=1,
        border_radius=15,
        color="#007B83",
        label_style=ft.TextStyle(size=15, color=cor_textos, weight=ft.FontWeight.BOLD),
        label='Categoria',
        options=[
            ft.dropdown.Option('Alimento', text_style=ft.TextStyle(size=15, color=cor_textos, weight=ft.FontWeight.BOLD)),
            ft.dropdown.Option('Transporte', text_style=ft.TextStyle(size=15, color=cor_textos, weight=ft.FontWeight.BOLD)), 
            ft.dropdown.Option('Salário', text_style=ft.TextStyle(size=15, color=cor_textos, weight=ft.FontWeight.BOLD)),
            ft.dropdown.Option('Lazer', text_style=ft.TextStyle(size=15, color=cor_textos, weight=ft.FontWeight.BOLD)),
            ft.dropdown.Option('Moradia', text_style=ft.TextStyle(size=15, color=cor_textos, weight=ft.FontWeight.BOLD)),
            ft.dropdown.Option('Vestiuário', text_style=ft.TextStyle(size=15, color=cor_textos, weight=ft.FontWeight.BOLD)),
            ft.dropdown.Option('Esposte', text_style=ft.TextStyle(size=15, color=cor_textos, weight=ft.FontWeight.BOLD)),
            ft.dropdown.Option('Empréstimos', text_style=ft.TextStyle(size=15, color=cor_textos, weight=ft.FontWeight.BOLD)),  
            ft.dropdown.Option('Outros', text_style=ft.TextStyle(size=15, color=cor_textos, weight=ft.FontWeight.BOLD)),                 
        ]
    )
     
    descricao = ft.TextField(label='Descrição', label_style=ft.TextStyle(size=15, color=cor_textos, weight=ft.FontWeight.BOLD), border_width=1, border_radius=15, color=cor_textos, text_style=ft.TextStyle(size=15, weight=ft.FontWeight.BOLD))

    valor =  ft.TextField(label='Valor', label_style=ft.TextStyle(size=15, color=cor_textos, weight=ft.FontWeight.BOLD), border_width=1, border_radius=15, color=cor_textos, text_style=ft.TextStyle(size=15, weight=ft.FontWeight.BOLD))

    forma = ft.Dropdown(
        bgcolor=verde,
        border_width=1,
        color=cor_textos,
        border_radius=15,
        label_style=ft.TextStyle(size=15, color=cor_textos, weight=ft.FontWeight.BOLD),
        label='Forma de Transação',
        options=[
            ft.dropdown.Option('Dinheiro', text_style=ft.TextStyle(size=15, color=cor_textos, weight=ft.FontWeight.BOLD)),
            ft.dropdown.Option('Cartão', text_style=ft.TextStyle(size=15, color=cor_textos, weight=ft.FontWeight.BOLD)), 
            ft.dropdown.Option('Pix', text_style=ft.TextStyle(size=15, color=cor_textos, weight=ft.FontWeight.BOLD)),
            ft.dropdown.Option('Fiado', text_style=ft.TextStyle(size=15, color=cor_textos, weight=ft.FontWeight.BOLD)), 
            ft.dropdown.Option('Outro', text_style=ft.TextStyle(size=15, color=cor_textos, weight=ft.FontWeight.BOLD)),                   
        ],
    )

    anom = ft.TextField(label='Ano',label_style=ft.TextStyle(size=15, color=cor_textos, weight=ft.FontWeight.BOLD), border_width=1, border_radius=15, color=cor_textos, text_style=ft.TextStyle(size=15, weight=ft.FontWeight.BOLD))
    mesm = ft.TextField(label='Mês',label_style=ft.TextStyle(size=15, color=cor_textos, weight=ft.FontWeight.BOLD), border_width=1, border_radius=15, color=cor_textos, text_style=ft.TextStyle(size=15, weight=ft.FontWeight.BOLD))
    diam = ft.TextField(label='Dia',label_style=ft.TextStyle(size=15, color=cor_textos, weight=ft.FontWeight.BOLD), border_width=1, border_radius=15, color=cor_textos, text_style=ft.TextStyle(size=15, weight=ft.FontWeight.BOLD))
   
    data_manual = ft.Container(
        content=ft.Column(
            [
                ft.Divider(),
                ft.Text(value='Ano, Mês e Dia para períodos passados', size=15, italic=True, color=cor_textos),
                anom,
                mesm,
                diam,
                ft.Divider(),
            ]
        )
    )

    historico = ft.Container(
        expand=True,
        padding = 10,
        margin = 5,
        content = ft.Column(
            [],
            scroll=ft.ScrollMode.AUTO
        )
    )


    # adicionar o alerta ao overlay
    def adicionar_alerta(alerta):
        if alerta not in page.overlay:
            page.overlay.append(alerta)
        alerta.open = True
        page.update()

    # remover o alerta ao overlay
    def remover_alerta(alerta):
        alerta.open = False
        page.update()
    
    def mostrar_alerta_erro_descricao():
            alerte_erro = ft.AlertDialog(
                title=ft.Text("Presta Atenção Abestado!", color=grafite),
                content=ft.Text('Descrição é obrigatório', color=vermelho),
                actions=[
                    ft.TextButton('Ok', on_click=lambda e: remover_alerta(alerte_erro)),
                ],
                actions_alignment=ft.MainAxisAlignment.END,
                open=True
            )
            adicionar_alerta(alerte_erro)

    def mostrar_alerta_erro_valor():
            alerte_erro = ft.AlertDialog(
                title=ft.Text("Presta Atenção Abestado!", color=grafite),
                content=ft.Text('Valor é obrigatório', color=vermelho),
                actions=[
                    ft.TextButton('Ok', on_click=lambda e: remover_alerta(alerte_erro)),
                ],
                actions_alignment=ft.MainAxisAlignment.END,
                open=True
            )
            adicionar_alerta(alerte_erro)
    
    def mostrar_alerta_erro_tipo():
            alerte_tipo = ft.AlertDialog(
            title=ft.Text("Presta Atenção Abestado!", color=grafite),
                content=ft.Text('Tipo é obrigatório', color=vermelho),
                actions=[
                    ft.TextButton('Ok', on_click=lambda e: remover_alerta(alerte_tipo)),
                ],
                actions_alignment=ft.MainAxisAlignment.END,
                open=True
            )
            adicionar_alerta(alerte_tipo)


    def salvar_dados(e): 

        if tipo.value is None or tipo.value == "":
            mostrar_alerta_erro_tipo()
            return

        if not descricao.value.strip():
            mostrar_alerta_erro_descricao()
            return
        try:
            valor_float = float(valor.value)
            if valor_float <= 0:
                raise ValueError("Valor deve ser maior que '0'!")
        except ValueError:
            mostrar_alerta_erro_valor()
            return 
   
        arquivo = "transacoes.xlsx"

        agora = datetime.now()

        if anom.value == "" and mesm.value == "" and diam.value == "":
            ano = agora.year
            mes = agora.month
            dia = agora.day
            hora = agora.strftime("%H:%M:%S")
        else:
            ano = anom.value if anom.value else agora.year
            mes = mesm.value if mesm.value else agora.month
            dia = diam.value if diam.value else agora.day
            hora = agora.strftime("%H:%M:%S") 
            
        # Verificando se o arquivo já existe
        if not os.path.exists(arquivo):
            # Cria um novo arquivo Excel e defino os cabeçalhos
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Transações"
            sheet.append(["Tipo", "Descrição", "Categoria", "Valor", "Forma de Transação", "Ano", "Mês", "Dia", "Hora"])
            workbook.save(arquivo)

        # Abrir o arquivo Excel para adicionar novos dados
        workbook = openpyxl.load_workbook(arquivo)
        sheet = workbook.active
        # Adicinar os dados do formulário ao Excel

        sheet.append([
            tipo.value,
            descricao.value,
            categoria.value,
            valor.value,
            forma.value,
            ano,
            mes,
            dia,
            hora
        ])
        # Salvar o arquivo
        workbook.save(arquivo)

        # Limpando os campos do formulário
        tipo.value = None
        descricao.value = " "
        categoria.value = None
        valor.value = " "
        forma.value = None

        anom.value = ""
        mesm.value = ""
        diam.value = ""
        
        # Atualiza o histórico assim que os dados forem salvos
        atualizar_historico()

        alerta_Form.open = False
        page.update()

    def atualizar_historico():
        # Limpando o histórico anterior

        historico.content.controls.clear()

        if os.path.exists("transacoes.xlsx"):
            workbook = openpyxl.load_workbook("transacoes.xlsx")
            sheet = workbook.active

            # iterar sobre as linhas do excel, começando da segunda linha
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Cria um novo container para cada transação

                # definir a cor da borda com base no tipo de transição
                tipo = row[0]
                if tipo == "Entrada":
                    cor = azul
                elif tipo == "Saída":
                    cor = vermelho
                else:
                    cor = grafite
                # Criando container para cada transação   
                trasacao = ft.Container(
                    border=ft.Border(left=ft.BorderSide(width=4, color=cor)),
                    margin=2,
                    padding=10,
                    border_radius=0,
                    content=ft.Row(
                        [
                            ft.Text(row[1], size=14, color=cor_textos, weight=ft.FontWeight.W_500),
                            ft.Text(f"{row[7]}/{row[6]}/{row[5]}", size=14, color=cor_textos, weight=ft.FontWeight.W_500),
                            ft.Text(f"R$ {row[3]}",  size=14, color=cor_textos, weight=ft.FontWeight.W_500),
                        ],
                        alignment=ft.MainAxisAlignment.SPACE_AROUND,
                        spacing=10,
                    )
                )
                # adiconar o novo container ao container de histórico
                historico.content.controls.append(trasacao)    
        atualizar_saldos()

    def atualizar_saldos():

        # Verificando se o arquivo já existe
        if not os.path.exists("transacoes.xlsx"):
            # Cria um novo arquivo Excel e defino os cabeçalhos
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Transações"
            sheet.append(["Tipo", "Descrição", "Categoria", "Valor", "Forma de Transação", "Ano", "Mês", "Dia", "Hora"])
            workbook.save("transacoes.xlsx")
        

        workbook = openpyxl.load_workbook("transacoes.xlsx")
        sheet = workbook.active
        en = 0
        sa = 0
            
        for row in sheet.iter_rows(min_row=2, values_only=True):
            valor = float(row[3])
            if row[0] == 'Entrada':
                en += valor
            elif row[0] == 'Saída':
                sa += valor
        to = en - sa

        total_entrada.content.controls[1].value = f"R$ {en:.2f}"
        total_saida.content.controls[1].value = f"R$ {sa:.2f}"
        saldo_total.content.controls[0].value =  f"R$ {to:.2f}"

        page.update()

    alerta_Form = ft.AlertDialog(
        shape=ft.RoundedRectangleBorder(radius=10),
        bgcolor=verde,
        content_padding=10,
        title=ft.Text(value='Nova transação', color=cor_textos),
        content=ft.Column(
            [
                tipo, 
                categoria,
                forma,
                descricao,
                valor,
                ft.Text(value='      A data será registrada automáticamente', color=cor_textos, size=12, italic=True),
                data_manual,
            ]
        ),
        actions=[
            ft.ElevatedButton('Salvar', on_click=salvar_dados, bgcolor="#007B83", color=cor_textos)
        ],
        open=False
    )


    # Associando o alerta a page
    page.overlay.append(alerta_Form)
    page.update()

    # Abrir alerta do formulário
    def formulario(e):
        alerta_Form.open = True
        page.update()
    
    def limpardados(e):
        historico.content.controls.clear() # Limpa o histórico da interface
        arquivo = "transacoes.xlsx" # Limpa o conteúdo do xlsx
        if os.path.exists(arquivo):
            workbook = openpyxl.load_workbook(arquivo)
            sheet = workbook.active

            # Manter o cabeçalho e apagar as outras linhas
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                for cell in row:
                    cell.value = None
            workbook.save(arquivo)

        total_entrada.content.controls[0].value = "R$ 0.00"
        total_saida.content.controls[0].value = "R$ 0.00"
        saldo_total.content.controls[1].value = "R$ 0.00"

        page.update()
    
    def mostrar_alerta_confirmacao(e):
        # Criar um Alerta
        alerta_confirmacao_limpeza = ft.AlertDialog(
            title=ft.Text("Confirmar Limpeza de dados"),
            content=ft.Text("Você tem certeza que deseja apagar todos os dados? esta ação é irreversível", color=vermelho, size=15, weight=ft.FontWeight.BOLD, italic=True),
            actions=[
                ft.TextButton("Cancelar", on_click=lambda e: remover_alerta(alerta_confirmacao_limpeza)),
                ft.ElevatedButton("Confirmar", bgcolor="#007B83", on_click=lambda e: [remover_alerta(alerta_confirmacao_limpeza),
                                                                   limpardados(e)
                ]
                ),
            ],
            actions_alignment=ft.MainAxisAlignment.SPACE_AROUND,
            open=True
        )
        adicionar_alerta(alerta_confirmacao_limpeza)
        atualizar_saldos()

    def adicionar_alerta(alerta):
        if alerta not in page.overlay:
            page.overlay.append(alerta)
        alerta.open = True
        page.update()  

    def remover_alerta(alerta):
        alerta.open = False
        page.update()
    
    def abrir_pg_analise(e):
        page.clean()
        page.add(pg_analise)
        page.update()
        
    def fecha_pg(e):
        page.clean()
        page.add(
        layout,
        botao_formulario
    )
        page.update()

    def calcular_totais_por_forma(df):
    # Verifica se o DataFrame está vazio
        if df.empty:
            return []

        # Agrupa por "Forma de Pagamento" e calcula o total
        totais = df.groupby('Forma de Transação')['Valor'].sum().reset_index()

        # Calcula o total geral
        total_geral = totais['Valor'].sum()

        # Adiciona uma coluna para calcular o percentual
        totais['Percentual'] = (totais['Valor'] / total_geral * 100).round(2)

        return totais


    def atualizar_descricao_forma(totais):
    # Limpa os controles da coluna antes de adicionar os novos
        descricao_forma.content.controls.clear()

        for _, row in totais.iterrows():
            # Cria os textos dinamicamente
            F_r = ft.Text(value=row['Forma de Transação'], size=18, color=cor_textos, weight=ft.FontWeight.W_500)
            F_v = ft.Text(value=f"R$ {row['Valor']:.2f}", size=18, color=cor_textos, weight=ft.FontWeight.W_500)
            F_p = ft.Text(value=f"{row['Percentual']} %", size=18, color=cor_textos, weight=ft.FontWeight.W_500)

            # Adiciona o container para cada forma de pagamento
            descricao_forma.content.controls.append(
                ft.Container(
                    margin=0,
                    padding=5,
                    border_radius=5,
                    expand=True,
                    gradient=ft.LinearGradient(
                        colors=[
                            "#007B83",
                            "#008A7A",
                            "#009570",
                            "#00A067",
                            "#00A86B",
                        ]
                    ),
                    content=ft.Row(
                        [
                            F_r,
                            F_v,
                            F_p
                        ],
                        expand=True, 
                        alignment=ft.MainAxisAlignment.SPACE_BETWEEN
                    )
                )
            )
        descricao_forma.update()  # Atualiza a interface

 
    data_inicial = ft.Text(value='Data Inicial', size=18, weight=ft.FontWeight.W_500, color="#A6A6A6")
    data_final = ft.Text(value='Data Final', size=18, weight=ft.FontWeight.W_500, color="#A6A6A6")

    def calcular_totais(df_filtrado):
        # Filtra todas as transações do tipo "Entrada"
        entradas = df_filtrado[df_filtrado['Tipo'] == 'Entrada']
        soma_entradas = entradas['Valor'].sum()
        qtd_entradas = len(entradas)

        # Filtra todas as transações do tipo "Saída"
        saidas = df_filtrado[df_filtrado['Tipo'] == 'Saída']
        soma_saidas = saidas['Valor'].sum()
        qtd_saidas = len(saidas)

    # Total geral de transações
        total_transacoes = soma_entradas + soma_saidas
        qtd_transacoes = qtd_entradas + qtd_saidas

    # Retornar os resultados
        return {
            'total_entradas': soma_entradas,
            'qtd_entradas': qtd_entradas,
            'total_saidas': soma_saidas,
            'qtd_saidas': qtd_saidas,
            'total_transacoes': total_transacoes,
            'qtd_transacoes': qtd_transacoes
        }
    

    def on_date_selected(e):

        global data_inicial_datetime, data_final_datetime

        selected_date = e.control.value

        data_formatada = selected_date.strftime("%d/%m/%y")
        if e.control.data == "from_date":
            data_inicial.value = f" De: {data_formatada}"
            data_inicial_datetime = selected_date
            data_inicial.update()
        elif e.control.data == "to_date":
            data_final.value = f"Até: {data_formatada} "
            data_final_datetime = selected_date
            data_final.update()

        # Chamar a função de filtragem apenas quando ambas as datas forem selecionadas
        # Checar se ambas as datas foram selecionadas (ou se data final foi preenchida automaticamente)
            while data_inicial_datetime is not None and data_final_datetime is not None:
                
                df_filtrados = filtrar_dados_por_periodo(data_inicial_datetime, data_final_datetime)
                resultados = calcular_totais(df_filtrados)

                # Usando os resultados nas variáveis
                quantidade_entrada.value = f"{resultados['qtd_entradas']}. Entradas"
                valor_entrada.value = f"R$      {resultados['total_entradas']:.2f}"
                quantidade_saida.value = f"{resultados['qtd_saidas']}. Saídas"
                valor_saida.value = f"R$      {resultados['total_saidas']:.2f}"
                quantidade_transacoes.value = f"{resultados['qtd_transacoes']}. Transações"
                valor_transacoes.value = f"R$      {resultados['total_transacoes']:.2f}"
                page.update()
                return resultados
            else:
                return
    

    #Aqui começa o tratamento para exibir informações sobre das transações e seus valores
    # dentro do período selecionado e filtrado por entrada ou saída
    def filtrando_tipo(e):
        global data_inicial_datetime, data_final_datetime, tipo_selecao
        # Verificar se as datas inicial e final foram selecionadas
        if data_inicial_datetime is None or data_final_datetime is None :
            snack_bar = ft.SnackBar(content=ft.Text("Por favor, selecione o período inicial e final.", color=ft.colors.BLACK), bgcolor=ft.colors.WHITE)
            page.overlay.append(snack_bar)
            snack_bar.open = True
            page.update()  # Atualiza para exibir o SnackBar
            return
         # Verificar se a data inicial é posterior à data final
        if data_inicial_datetime > data_final_datetime:
            snack_bar = ft.SnackBar(content=ft.Text("A data inicial não pode ser posterior à data final."))
            page.overlay.append(snack_bar)
            snack_bar.open = True
            page.update()
            return
        tipo = e.control.data

        # Checar se ambas as datas foram selecionadas (ou se data final foi preenchida automaticamente
        if data_inicial_datetime is not None and data_final_datetime is not None:
            df = filtrar_dados_por_periodo(data_inicial_datetime, data_final_datetime)
            if tipo == 'E':
                tipo_selecao = 'ENTRADAS'
                rotulo_forma.value = f'   Formas de {tipo_selecao} :' 
                rotulo_forma.update()
                categoria_rotulo.value = f'   {tipo_selecao} entre as categorias'
                categoria_rotulo.update()
                rotulo_resumo.value = f'   Resumo das {tipo_selecao}'
                rotulo_resumo.update()
                df_entradas = df[df['Tipo'] == 'Entrada']
                if df_entradas.empty:
                    snack_bar = ft.SnackBar(content=ft.Text("Não há registros de entradas no período selecionado!", color=ft.colors.BLACK), bgcolor=ft.colors.WHITE)
                    page.overlay.append(snack_bar)
                    snack_bar.open = True
                    page.update()
                    snack_bar.open()
                    return
                categorias_metricas = calcular_metricas_por_categoria(df_entradas)
                atualizar_categorias_na_interface(categorias_metricas, coluna_categorias)
                df_entradas_processado = processa_dados(df_entradas)
                # Processa os totais por forma de pagamento
                totais = calcular_totais_por_forma(df_entradas)
                # Atualiza os componentes
                atualizar_descricao_forma(totais)
                return df_entradas_processado
            elif tipo == 'S':
                tipo_selecao = 'SAÍDAS'
                rotulo_forma.value = f'   Formas de {tipo_selecao}:'
                rotulo_forma.update()
                categoria_rotulo.value = f'   {tipo_selecao} entre as categorias:'
                categoria_rotulo.update()
                rotulo_resumo.value = f'   Resumo das {tipo_selecao}'
                rotulo_resumo.update()
                df_saidas = df[df['Tipo'] == 'Saída']
                if df_saidas.empty:
                    snack_bar = ft.SnackBar(content=ft.Text("Não há registros de saídas no período selecionado!", color=ft.colors.BLACK),bgcolor=ft.colors.WHITE)
                    page.overlay.append(snack_bar)
                    snack_bar.open = True
                    page.update()
                    snack_bar.open()
                    return
                # Processa as métricas por categoria
                categorias_metricas = calcular_metricas_por_categoria(df_saidas)

            # Atualiza a interface com os dados processados
                atualizar_categorias_na_interface(categorias_metricas, coluna_categorias)
                df_saidas_processado = processa_dados(df_saidas)
                # Processa os totais por forma de pagamento
                totais = calcular_totais_por_forma(df_saidas)
                # Atualiza os componentes
                atualizar_descricao_forma(totais)
                return df_saidas_processado
        else:
            return
        
    coluna_categorias = ft.Row(
    controls=[],  # Inicialmente vazio
    scroll=ft.ScrollMode.AUTO,  # Permite rolar caso os itens ultrapassem o espaço visível
    alignment=ft.MainAxisAlignment.START,
    spacing=10,
)

    lista_trasacoes = []
    
    def processa_dados(df):

        # 1º calcular o valor total
        total_valor = df['Valor'].sum()
        # 2. Agrupe o DataFrame por 'Descrição' para calcular a quantidade e o valor total de cada grupo
        agrupamento = df.groupby('Descrição').agg(
            quantidade=('Valor', 'size'),        # Conta quantas vezes a descrição aparece
            valor_total=('Valor', 'sum')         # Soma os valores para cada descrição
        ).reset_index()  # Convertemos para um DataFrame padrão após o agrupamento    
        # 3. Calcule a porcentagem de cada descrição em relação ao total do período selecionado
        agrupamento['percentual'] = (agrupamento['valor_total'] / total_valor * 100).round(2)

        def listas_dados_processado(df):
            global lista_trasacoes 

            lista_processada = df.values.tolist()
            lista_trasacoes = lista_processada
            return lista_processada

        listas_processadas = listas_dados_processado(agrupamento)

        desc_porc_real.content.controls.clear()

        for i in listas_processadas:
            trasa_text = f"{i[1]} x {i[0]}  R$ {i[2]:.2f}  total de {i[3]} %"
            u = ft.Row([ft.Text(trasa_text, style=ft.TextStyle(size=15, color=cor_textos))])
            # Adiciona a linha ao container desc_porc_real
            desc_porc_real.content.controls.append(u)
        # Atualiza o container para exibir as novas linhas
        desc_porc_real.update()
        


    # Função que usa os objetos datetime
    def filtrar_dados_por_periodo(data_inicial_datetime, data_final_datetime):
    # Carrega o arquivo Excel
        df = pd.read_excel("transacoes.xlsx")

    # Criar a coluna de data no formato datetime
        df['Data'] = pd.to_datetime(df[['Ano', 'Mês', 'Dia']].rename(columns={'Ano': 'year', 'Mês': 'month', 'Dia': 'day'}))

    # Filtrar o dataframe pelo período selecionado
        df_filtrado = df[(df['Data'] >= data_inicial_datetime) & (df['Data'] <= data_final_datetime)]

        return df_filtrado

    datepicker_de = ft.DatePicker(
        cancel_text='Cancelar',
        confirm_text='confirmar',
        field_hint_text='MM/DD/AAAA',
        field_label_text='Selecione uma Data',
        help_text='Calendário',
        open=False,
        data="from_date",
        on_change=on_date_selected 
    )

    datepicker_ate = ft.DatePicker(
        cancel_text='Cancelar',
        confirm_text='confirmar',
        field_hint_text='MM/DD/AAAA',
        field_label_text='Selecione uma Data',
        help_text='Calendário',
        open=False,
        data="to_date",
        on_change=on_date_selected 
    )

    def abrir_date_de(e):
        e.page.overlay.append(datepicker_de)
        datepicker_de.open = True
        e.page.update()

    def abrir_date_ate(e):
        e.page.overlay.append(datepicker_ate)
        datepicker_ate.open = True
        e.page.update()

    quantidade_entrada = ft.Text(value="", color=cor_textos, size=18, weight=ft.FontWeight.W_500)
    valor_entrada = ft.Text(value="", color=cor_textos, size=18, weight=ft.FontWeight.W_500)
    quantidade_saida = ft.Text(value="", color=cor_textos, size=18, weight=ft.FontWeight.W_500)
    valor_saida = ft.Text(value="", color=cor_textos, size=18, weight=ft.FontWeight.W_500)
    quantidade_transacoes = ft.Text(value="", color=cor_textos, size=18, weight=ft.FontWeight.W_500)
    valor_transacoes = ft.Text(value="", color=cor_textos, size=18, weight=ft.FontWeight.W_500)

    
    btn_entrada = ft.ElevatedButton(text='ENTRADA',  color=cor_textos, data='E', on_click=filtrando_tipo, bgcolor="#007B83", elevation=0.8)
    btn_saida = ft.ElevatedButton(text='SAÍDA',  color=cor_textos, data='S', on_click=filtrando_tipo, bgcolor="#007B83", elevation=0.8)
  

    filtro_tipo = ft.Container(
        margin=10,
        padding=10,
        border_radius=20,
        bgcolor=preto,
        content=ft.Row(
            [
                btn_entrada,
                btn_saida
            ],
            alignment=ft.MainAxisAlignment.SPACE_AROUND
        )
    )

    infor_geral = ft.Container(
        padding=10,
        margin=15,
        border=ft.border.all(width=0.5, color=ft.colors.WHITE),
        gradient=ft.LinearGradient(
            colors=[
                "#007B83",
                "#008A7A",
                "#009570",
                "#00A067",
                "#00A86B",
                ]
            ),
        border_radius=10,
        content=ft.Column(
            [
                ft.Row(
                    [
                        quantidade_entrada,                      
                        valor_entrada,
                    ],
                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN
                ),
                ft.Row(
                    [
                        quantidade_saida,
                        valor_saida,
                    ],
                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN
                ),
                ft.Row(
                    [
                        quantidade_transacoes,
                        valor_transacoes,
                    ],
                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN
                )
            ]
        )
    )

    F_r = ft.Text(value='')
    F_v = ft.Text(value='')
    F_p = ft.Text(value='')
   

    descricao_forma = ft.Container(
        border_radius=10,
        margin=10,
        expand=True,
        content=ft.Column(
            controls=[
                ft.Container(
                    expand=True,
                    border_radius=20,
                    margin=2, 
                    content=ft.Row(
                        [
                            F_r,
                            F_v,
                            F_p
                        ],
                        expand=True, alignment=ft.MainAxisAlignment.SPACE_EVENLY
                    )
                ),
            ]
        )
    ) 


    def calcular_metricas_por_categoria(df):
        """
        Calcula métricas (quantidade, soma, média, mínimo, máximo) para cada categoria no DataFrame.
        
        :param df: DataFrame filtrado contendo os dados
        :return: DataFrame com as métricas calculadas
        """
       
        categorias = df.groupby("Categoria").agg(
            Quantidade=("Valor", "count"),
            Soma=("Valor", "sum"),
            Média=("Valor", "mean"),
            Mínimo=("Valor", "min"),
            Máximo=("Valor", "max")
        ).reset_index()
        return categorias
    
    def atualizar_categorias_na_interface(categorias_df, coluna_categorias):
        """
        Atualiza os componentes da interface para refletir os dados processados por categoria.
        
        :param categorias_df: DataFrame com as métricas por categoria
        :param coluna_categorias: Componente ft.Row contendo os contêineres das categorias
        """
        # Limpa os componentes antigos
        coluna_categorias.controls.clear()

        # Itera sobre as linhas do DataFrame para criar novos componentes
        for _, row in categorias_df.iterrows():
            nome_cat = ft.Text(f"{row['Categoria']}", size=20, color=cor_textos, weight=ft.FontWeight.BOLD)
            qua_cat = ft.Text(f"Quantidade: {row['Quantidade']}", size=14, color=cor_textos, weight=ft.FontWeight.W_500)
            val_cat = ft.Text(f"Valor: R$ {row['Soma']:.2f}", size=14, color=cor_textos, weight=ft.FontWeight.W_500)
            por_cat = ft.Text(f"Porcent: {row['Soma'] / categorias_df['Soma'].sum() * 100:.1f}%", size=14, color=cor_textos, weight=ft.FontWeight.W_500)
            vme_cat = ft.Text(f"Valor Méd: R$ {row['Média']:.2f}", size=14, color=cor_textos, weight=ft.FontWeight.W_500)
            vma_cat = ft.Text(f"Valor Máx: R$ {row['Máximo']:.2f}", size=14, color=cor_textos, weight=ft.FontWeight.W_500)
            vmi_cat = ft.Text(f"Valor Mín: R$ {row['Mínimo']:.2f}", size=14, color=cor_textos, weight=ft.FontWeight.W_500)

            categoria_resumo = ft.Container(
                gradient=ft.LinearGradient(
                    colors=[
                        "#007B83",
                        "#008A7A",
                        "#009570",
                        "#00A067",
                        "#00A86B",
                    ]
                ),
                border_radius=20,
                height=250,
                width=150,
                margin=15,
                shadow=ft.BoxShadow(spread_radius=0.5, blur_radius=1.2, color=ft.colors.WHITE),
                content=ft.Column(
                    controls=[nome_cat, qua_cat, val_cat, por_cat, vme_cat, vma_cat, vmi_cat],
                    alignment=ft.MainAxisAlignment.START,
                    horizontal_alignment=ft.CrossAxisAlignment.CENTER
                )
            )

            # Adiciona o contêiner à linha
            coluna_categorias.controls.append(categoria_resumo)

        # Atualiza a interface
        coluna_categorias.update()

    

      

    rotulo_forma = ft.Text(value=f'Antes de mais nada, defina um período com uma data inicial e uma data final, para que sejam feitas as análises sobre as transações. Depois click em entradas e/ou saídas então você verá como elas foram divididas entre as formas, como pix, dinheiro etc. Bem como a distribuição entre as categorias como alimentação, lazer entre outros e um resumo das transações.     Faça um bom Uso.', color="#A6A6A6", size=15, weight=ft.FontWeight.W_500, text_align=ft.TextAlign.JUSTIFY)
    categoria_rotulo = ft.Text(value='', size=15, color="#A6A6A6", weight=ft.FontWeight.W_500)
    rotulo_resumo = ft.Text(value=' ', size=15, color="#A6A6A6", weight=ft.FontWeight.W_500)

    painel = ft.Container(
        padding=10,
        margin=5,
        content=ft.Column(
            [                
                rotulo_forma,
                ft.Row(
                    [
                        descricao_forma
                    ],
                    alignment=ft.MainAxisAlignment.SPACE_AROUND,
                    spacing=0.5
                ),
                categoria_rotulo,
                coluna_categorias,
                rotulo_resumo,
                desc_porc_real
            ],
            alignment=ft.MainAxisAlignment.START,
            scroll=ft.ScrollMode.AUTO,
        ),
        expand=True,
    )


    pg_analise = ft.Container(
        expand=True,
        bgcolor=preto,
        padding=15,
        content=ft.Column(
            [
                ft.Row(
                    [
                        ft.IconButton(icon=ft.icons.CLOSE, icon_color=ft.colors.WHITE, icon_size=20, on_click=fecha_pg),
                    ],
                    alignment=ft.MainAxisAlignment.END,
                ),
                ft.Column(
                    [
                        ft.Row(
                            [
                                ft.Text(value=" Período", color="#A6A6A6", size=20, weight=ft.FontWeight.BOLD, expand=True, text_align=ft.alignment.center_left)
                            ],
                            alignment=ft.alignment.center_left
                        ),
                        ft.Container(
                            bgcolor=preto,
                            content=ft.Row(
                                [
                                data_inicial,
                                ft.IconButton(icon=ft.icons.CALENDAR_MONTH, icon_color="#A6A6A6", on_click=abrir_date_de),
                                data_final,
                                ft.IconButton(icon=ft.icons.CALENDAR_MONTH, icon_color="#A6A6A6", on_click=abrir_date_ate)
                                ],
                                alignment=ft.MainAxisAlignment.SPACE_AROUND
                            ),
                        )
                    ],
                    alignment=ft.MainAxisAlignment.SPACE_AROUND,
                ),
                infor_geral,
                filtro_tipo,
                ft.Divider(height=5, thickness=3, color=ft.colors.GREEN_100),
                painel,
            ]
        )
    )

    analise = ft.IconButton(icon=ft.icons.BAR_CHART_ROUNDED, icon_color=verde, icon_size=25, on_click=abrir_pg_analise)
    btn_limpardados = ft.IconButton(icon=ft.icons.DELETE_FOREVER, icon_color=vermelho, icon_size=25, on_click=mostrar_alerta_confirmacao)

    layout = ft.Container(
        expand=True,
        bgcolor= '#2C2C2E',
        border_radius=5,
        padding=5,
        content=ft.Column(
            [
                ft.Row([analise], alignment=ft.MainAxisAlignment.END),
                ft.Container(
                    margin=15,
                    border=ft.border.all(width=0.5, color=ft.colors.WHITE),
                    gradient=ft.LinearGradient(
                        colors=[
                            "#007B83",
                            "#008A7A",
                            "#009570",
                            "#00A067",
                            "#00A86B",
                        ]
                    ),
                    border_radius=10,
                    content=ft.Column(
                        [
                            ft.Row([saldo_total], alignment=ft.MainAxisAlignment.END),
                            ft.Row([total_entrada, total_saida], alignment=ft.MainAxisAlignment.CENTER, expand=True),
                        ]
                    )
                ),
                ft.Row(
                    [
                        ft.Text(value='HISTÓRICO', size=15, weight=ft.FontWeight.BOLD, color="#A6A6A6"),
                        btn_limpardados,
                    ],
                    alignment=ft.MainAxisAlignment.SPACE_AROUND,
                ),
                historico 
            ],
            spacing=10,
        )
    )

    botao_formulario =  ft.FloatingActionButton(

            elevation=10,
            bgcolor=verde,
            shape=ft.CircleBorder(),
            icon=ft.icons.ADD, 
            on_click=formulario)


    # Inicia o app buscando o histórico atualizado
    atualizar_historico()

    page.add(
        layout,
        botao_formulario,
    )
if __name__ == "__main__":
    ft.app(target=main, view=ft.AppView.FLET_APP)